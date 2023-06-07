import os
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_folder_path():
    script_directory = os.path.dirname(os.path.abspath(__file__))
    folder_name = "png"
    return os.path.join(script_directory, folder_name)


def get_png_files(folder_path):
    return [file for file in os.listdir(folder_path) if file.endswith(".png")]


def get_file_time(png_files):
    file_time = [file[8:-9].replace('_', ':') if file.endswith("_Full.png") else file for file in png_files]
    return file_time


def define_time(file_time):
    return ['PAGI' if '06:00:00' <= t <= '18:00:00' else 'MALAM' for t in file_time]


def create_excel_file():
    workbook = Workbook()
    return workbook, workbook.active


def set_merged_title(sheet):
    merged_value = input("Masukan Judul Excel: ")
    sheet.merge_cells("A1:J2")
    sheet["A1"].value = merged_value
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")


def set_header_values(sheet):
    header_values = [
        "Nomor",
        "Nama Foto",
        "Bacaan Plat Kiri",
        "Pelat Tertutup Kirim",
        "Pelat Tidak Jelas Kiri",
        "Bacaan Pelat Kanan",
        "Pelat Tertutup Kanan",
        "Pelat Tidak Jelas Kanan",
        "Xml ADA / TIDAK_ADA",
        "Jam PAGI / MALAM"
    ]
    for col, value in enumerate(header_values, start=1):
        column_letter = get_column_letter(col)
        sheet[column_letter + "3"].value = value
    return len(header_values)


def set_column_width(sheet, column_count, column_width=25):
    for col in range(1, column_count + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = column_width


def set_border(sheet):
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in sheet.iter_rows(min_row=1, max_row=3, max_col=sheet.max_column):
        for cell in row:
            cell.border = border


def populate_data(sheet, png_files, defineTime):
    for index, file_name in enumerate(png_files, start=1):
        number = index
        number += 0
        sheet.cell(row=index + 3, column=1, value=number)
        sheet.cell(row=index + 3, column=2, value=file_name)
        sheet.cell(row=index + 3, column=10, value=defineTime[index - 1])


def set_row_border(sheet):
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.row == 4:
                cell.border = border


def export_excel_file(workbook):
    print("Loading Slur...")
    time.sleep(3)
    excel_file_name = workbook["Sheet"].cell(row=1, column=1).value
    workbook.save(f"{excel_file_name}.xlsx")
    print(f"Excel file '{excel_file_name}' successfully exported.")


def main():
    folder_path = get_folder_path()
    png_files = get_png_files(folder_path)
    file_time = get_file_time(png_files)
    defineTime = define_time(file_time)

    workbook, sheet = create_excel_file()
    set_merged_title(sheet)
    column_count = set_header_values(sheet)
    set_column_width(sheet, column_count)
    set_border(sheet)
    populate_data(sheet, png_files, defineTime)
    set_row_border(sheet)
    export_excel_file(workbook)


if __name__ == '__main__':
    main()